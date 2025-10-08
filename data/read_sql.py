from openpyxl import load_workbook

# Replace 'sample.xlsx' with your file's name or path
filename = 'sql.xlsx'

try:
    # Load the workbook from the file
    workbook = load_workbook(filename=filename)
    
    # Get the active worksheet
    sheet = workbook.active
    print(f"Successfully loaded worksheet: '{sheet.title}'")
    print("\n--- Reading all data ---")
    for row in sheet.iter_rows(values_only=True):
        print(row)
except FileNotFoundError:
    print(f"Error: The file '{filename}' was not found.")
except Exception as e:
    print(f"An error occurred: {e}")


